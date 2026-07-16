import sys, pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding="utf-8")

SRC = r"C:\Users\praks\OneDrive\Desktop\Rec app.xlsx"
OUT = r"C:\Users\praks\Music\RecTrack-UX\RecTrack_Migration_Simple.xlsx"

STAGE = "Not Started,In Progress,Blocked,Completed"
ROLE  = "Program Head,Trainer,Editor,Reviewer,Manager"
TTYPE = "Internal,External"
STUDIO= "Studio 1,Studio 2,Studio 3,Studio 4"
SLOTST= "Scheduled,In Progress,Completed,Missed,Rescheduled"
REC   = "Recording Started,Recording In Progress,Recording Completed"
REASON= "Given other training,No-show,Late start,Swapped by others,Tech issue,Trainer on leave"
YN    = "Yes,No"

ARIAL="Arial"
HF=PatternFill("solid",fgColor="3B3F8C"); HINT=PatternFill("solid",fgColor="EEF0FF")
CALC=PatternFill("solid",fgColor="FFF7E0"); KEYINFO=PatternFill("solid",fgColor="E6F6EC")
HFONT=Font(name=ARIAL,bold=True,color="FFFFFF",size=10); BASE=Font(name=ARIAL,size=10)
TITLE=Font(name=ARIAL,bold=True,size=15,color="1F2340"); SUB=Font(name=ARIAL,bold=True,size=11,color="3B3F8C")
thin=Side(style="thin",color="DADDE3"); BD=Border(left=thin,right=thin,top=thin,bottom=thin)

def sheet(wb,name,headers,rows,widths=None,dvs=None,hint=None,calc_cols=()):
    ws=wb.create_sheet(name); r0=1
    if hint:
        c=ws.cell(1,1,hint); c.font=Font(name=ARIAL,italic=True,size=9,color="3B3F8C"); c.fill=HINT; r0=2
    for j,h in enumerate(headers,1):
        c=ws.cell(r0,j,h); c.font=HFONT; c.fill=HF; c.alignment=Alignment("left","center"); c.border=BD
    for i,row in enumerate(rows,r0+1):
        for j,val in enumerate(row,1):
            c=ws.cell(i,j,val); c.font=BASE; c.border=BD; c.alignment=Alignment(vertical="center")
            if headers[j-1] in calc_cols: c.fill=CALC
    for j,h in enumerate(headers,1):
        ws.column_dimensions[get_column_letter(j)].width=(widths or {}).get(h,max(11,min(34,len(str(h))+3)))
    ws.freeze_panes=ws.cell(r0+1,1); ws.row_dimensions[r0].height=20
    if dvs:
        for h,lst in dvs.items():
            if h in headers:
                col=get_column_letter(headers.index(h)+1)
                dv=DataValidation(type="list",formula1=f'"{lst}"',allow_blank=True)
                ws.add_data_validation(dv); dv.add(f"{col}{r0+1}:{col}{r0+3000}")
    return ws,r0

# ---- read your Excel, clean hierarchy, one row per sub-topic ----
df=pd.read_excel(SRC,sheet_name="Subjects and Status")
df.columns=[str(c).strip() for c in df.columns]
for c in ["Main Subject","Subject","ChapterName","TopicName"]: df[c]=df[c].ffill()
df["TopicName"]=df["TopicName"].fillna("General")
df=df[df["Sub-topics"].notna()].copy().reset_index(drop=True)

rows=[]
for idx,r in df.iterrows():
    f = idx==0
    rows.append([
        r["Main Subject"], r["Subject"], r["ChapterName"], r["TopicName"], str(r["Sub-topics"]).strip(),
        "Madhu" if f else "", "Internal",
        "2026-06-03" if f else "", "2026-06-05" if f else "",   # Start Date, Target End Date
        "Completed" if f else "", "Not Started" if f else "", "In Progress" if f else "",
        "", "", "", "",        # Shooting, Shoot Review, Editing, Final Review
        "",                    # Complete % (formula filled after)
        ""                     # Notes
    ])

wb=Workbook(); wb.remove(wb.active)

# Read me
ws=wb.create_sheet("Read me"); ws.column_dimensions["A"].width=120; ws.sheet_view.showGridLines=False
for i,(t,k) in enumerate([
  ("RecTrack — Simple Migration File","title"),
  ("3 working tabs. Plain language only — no ID numbers, ever.","sub"),
  ("",""),
  ("• 'Curriculum + Status' = main tab. One row per sub-topic. Start Date + Target End Date, then pick each stage's status.","n"),
  ("• 'Complete %' fills in AUTOMATICALLY from the stage statuses × their weights — you never type a percentage.","calc"),
  ("   weights: PPT 20% · Script 20% · Presentation 15% · Shooting 25% · Editing 10% · Final Review 10%  (Shoot Review = gate).","n"),
  ("• 'People' = team + role.   • 'Schedule' = recording slots + WHY a slot was missed (the defaulter reasons).","n"),
  ("",""),
  ("NO ID COLUMNS — the system creates all keys on import, so the 'duplicate IDs from copies' problem can't return.","keyinfo"),
  ("",""),
  ("Core Java is filled in from your file as the example.","n"),
], 1):
    fonts={"title":TITLE,"sub":SUB,"calc":Font(name=ARIAL,size=10,bold=True,color="8A6D00"),
           "keyinfo":Font(name=ARIAL,size=10,bold=True,color="0F7A37")}
    c=ws.cell(i,1,t); c.font=fonts.get(k,BASE)
    if k=="calc": c.fill=CALC
    if k=="keyinfo": c.fill=KEYINFO

# 1. Curriculum + Status
heads=["Main Subject","Subject","Chapter","Topic","Sub-topic","Planned Trainer","Trainer Type",
       "Start Date","Target End Date","PPT","Script","Presentation","Shooting","Shooting Review",
       "Editing","Final Review","Complete %","Notes"]
ws,r0=sheet(wb,"Curriculum + Status",heads,rows,
   widths={"Main Subject":14,"Subject":12,"Chapter":20,"Topic":15,"Sub-topic":28,"Planned Trainer":13,
           "Trainer Type":11,"Start Date":12,"Target End Date":14,"Complete %":11,"Notes":18},
   dvs={"Trainer Type":TTYPE,"PPT":STAGE,"Script":STAGE,"Presentation":STAGE,"Shooting":STAGE,
        "Shooting Review":STAGE,"Editing":STAGE,"Final Review":STAGE},
   calc_cols={"Complete %"},
   hint="One row per sub-topic. Start + Target End date, then pick each stage's status. 'Complete %' auto-calculates (yellow).")
# inject the auto % formula (status × weight). cols: PPT=J Script=K Pres=L Shoot=M Edit=O Final=P ; Complete%=Q
for i in range(len(rows)):
    r=r0+1+i
    ws.cell(r,17).value=(f'=0.2*(J{r}="Completed")+0.2*(K{r}="Completed")+0.15*(L{r}="Completed")'
                         f'+0.25*(M{r}="Completed")+0.1*(O{r}="Completed")+0.1*(P{r}="Completed")')
    ws.cell(r,17).number_format="0%"; ws.cell(r,17).font=BASE; ws.cell(r,17).border=BD; ws.cell(r,17).fill=CALC

# 2. People
sheet(wb,"People",["Name","Email","Role","Internal/External","Active"],
   [["Madhu","madhu@example.com","Trainer","Internal","Yes"],
    ["Akash","akash@example.com","Trainer","Internal","Yes"],
    ["Leo","leo@example.com","Trainer","Internal","Yes"],
    ["Amy","amy@example.com","Trainer","Internal","Yes"],
    ["Pamy","pamy@example.com","Trainer","External","Yes"],
    ["Priya","priya@example.com","Program Head","","Yes"],
    ["(add editor later)","","Editor","","Yes"],
    ["(add reviewer later)","","Reviewer","","Yes"]],
   widths={"Email":26,"Name":18,"Internal/External":15},
   dvs={"Role":ROLE,"Internal/External":TTYPE,"Active":YN},
   hint="Phase 1 = Program Head + Trainers. Add Editor / Reviewer / Manager when you switch them on.")

# 3. Schedule  (room management + defaulter reasons)
sheet(wb,"Schedule",
   ["Date","Start","End","Studio","Trainer","Sub-topic","Status","Recording Status","Reason if Missed","Revised Date","Notes"],
   [["2026-06-03","10:30","12:30","Studio 1","Akash","Print","Completed","Recording Completed","","",""],
    ["2026-06-03","12:30","14:30","Studio 2","Leo","Operators","In Progress","Recording In Progress","","",""],
    ["2026-06-04","08:30","10:30","Studio 3","Amy","Inheritance","Scheduled","","","",""],
    ["2026-06-05","12:30","14:30","Studio 4","Pamy","Arrays","Missed","","Given other training","2026-06-06","pulled to a live batch"]],
   widths={"Sub-topic":20,"Reason if Missed":18,"Revised Date":12,"Notes":22},
   dvs={"Studio":STUDIO,"Status":SLOTST,"Recording Status":REC,"Reason if Missed":REASON},
   hint="Scarce rooms = this matters most. 'Reason if Missed' captures defaulters (given other training, swapped, no-show, late).")

wb.save(OUT)
print("SAVED:",OUT)
print("sub-topic rows:",len(rows),"| tabs:",wb.sheetnames)
