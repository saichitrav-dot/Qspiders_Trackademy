import sys, pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding="utf-8")

SRC = r"C:\Users\praks\OneDrive\Desktop\Rec app.xlsx"
OUT = r"C:\Users\praks\Music\RecTrack-UX\RecTrack_DataModel.xlsx"

# ---------- enums ----------
STAGE_STATUS = "Not Started,In Progress,Blocked,Completed"
SLOT_STATUS  = "Scheduled,In Progress,Completed,Missed,Rescheduled,Swapped"
REC_STATUS   = "Not Started,Recording,Completed"
EDIT_STATUS  = "Not Started,In Progress,Rework,Completed"
REVIEW_STATUS= "Pending,Approved,Rejected,Rework"
QUALITY      = "Excellent,Good,Average,Poor"
ROLE         = "admin,trainer,editor,reviewer,manager"
TTYPE        = "Internal,External"
YN           = "Yes,No"
REVIEW_TYPE  = "Shooting Review,Final Review"

# ---------- styling ----------
ARIAL   = "Arial"
HDR_FILL = PatternFill("solid", fgColor="3B3F8C")
ID_FILL  = PatternFill("solid", fgColor="E8EAF6")
FK_FILL  = PatternFill("solid", fgColor="F1F3F5")
NOTE_FILL= PatternFill("solid", fgColor="FFF7E0")
HDR_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
BASE_FONT= Font(name=ARIAL, size=10)
TITLE_FONT=Font(name=ARIAL, bold=True, size=14, color="1F2340")
SUB_FONT = Font(name=ARIAL, bold=True, size=11, color="3B3F8C")
thin = Side(style="thin", color="D9DCE3")
BORDER = Border(left=thin,right=thin,top=thin,bottom=thin)

def write_table(wb, name, headers, rows, id_cols=(), fk_cols=(), widths=None,
                validations=None, note=None):
    ws = wb.create_sheet(name)
    r0 = 1
    if note:
        ws.cell(1,1,note).font = Font(name=ARIAL, italic=True, size=9, color="8A6D00")
        ws.cell(1,1).fill = NOTE_FILL
        r0 = 2
    for j,h in enumerate(headers,1):
        c = ws.cell(r0,j,h); c.font=HDR_FONT; c.fill=HDR_FILL
        c.alignment=Alignment(horizontal="left",vertical="center"); c.border=BORDER
    for i,row in enumerate(rows, r0+1):
        for j,val in enumerate(row,1):
            c = ws.cell(i,j,val); c.font=BASE_FONT; c.border=BORDER
            c.alignment=Alignment(vertical="center")
            h = headers[j-1]
            if h in id_cols: c.fill=ID_FILL
            elif h in fk_cols: c.fill=FK_FILL
    # widths
    for j,h in enumerate(headers,1):
        w = (widths or {}).get(h, max(12, min(40, len(str(h))+4)))
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(r0+1,1)
    ws.row_dimensions[r0].height = 20
    # validations: dict header-> csv list
    if validations:
        for h, lst in validations.items():
            if h in headers:
                col = get_column_letter(headers.index(h)+1)
                dv = DataValidation(type="list", formula1=f'"{lst}"', allow_blank=True)
                ws.add_data_validation(dv)
                dv.add(f"{col}{r0+1}:{col}{r0+2000}")
    return ws

# ---------- migrate Core Java curriculum ----------
df = pd.read_excel(SRC, sheet_name="Subjects and Status")
df.columns = [str(c).strip() for c in df.columns]
for c in ["Main Subject","Subject","ChapterName","TopicName"]:
    df[c] = df[c].ffill()
df["TopicName"] = df["TopicName"].fillna("General")
df = df[df["Sub-topics"].notna()].copy().reset_index(drop=True)

main_rows, subj_rows, chap_rows, top_rows, st_rows, ci_rows = [],[],[],[],[],[]
main_map, subj_map, chap_map, top_map = {}, {}, {}, {}
mc=sc=cc=tc=stc=0
for _,r in df.iterrows():
    ms, su, ch, tp, sb = r["Main Subject"], r["Subject"], r["ChapterName"], r["TopicName"], str(r["Sub-topics"]).strip()
    if ms not in main_map:
        mc+=1; mid=f"MS{mc:02d}"; main_map[ms]=mid; main_rows.append([mid, ms])
    mid=main_map[ms]
    k=(mid,su)
    if k not in subj_map:
        sc+=1; sid=f"SUB{sc:03d}"; subj_map[k]=sid; subj_rows.append([sid, mid, su, sc])
    sid=subj_map[k]
    k=(sid,ch)
    if k not in chap_map:
        cc+=1; cid=f"CH{cc:04d}"; chap_map[k]=cid; chap_rows.append([cid, sid, ch, len(chap_rows)+1])
    cid=chap_map[k]
    k=(cid,tp)
    if k not in top_map:
        tc+=1; tid=f"TP{tc:04d}"; top_map[k]=tid; top_rows.append([tid, cid, tp, len(top_rows)+1])
    tid=top_map[k]
    stc+=1; stid=f"ST{stc:05d}"
    seq=sum(1 for x in st_rows if x[1]==tid)+1
    st_rows.append([stid, tid, sb, seq])
    ci_rows.append([f"CI{stc:05d}", stid, "P001", None, ""])  # planned trainer Madhu(P001)

# ---------- build workbook ----------
wb = Workbook(); wb.remove(wb.active)

# README
ws = wb.create_sheet("_README")
lines = [
 ("RecTrack — Data Model (normalised, ID-keyed)","title"),
 ("Replaces the flat Excel. Every entity has a permanent ID; names can change without breaking links.","sub"),
 ("","n"),
 ("LOCKED DECISIONS","sub"),
 ("• Deliverable = one CONTENT ITEM per Sub-topic (recording happens at sub-topic level).","n"),
 ("• Pipeline = 7 steps with TWO review gates: PPT → Script → Presentation → Shooting → Shooting Review → Editing → Final Review.","n"),
 ("• Scale = 100 Subjects, each with chapters/topics/sub-topics → thousands of content items. (Core Java migrated here as the example.)","n"),
 ("","n"),
 ("ENTITY MAP","sub"),
 ("main_subject → subject → chapter → topic → subtopic → content_item (the spine)","n"),
 ("content_item ─< item_stage (7 rows: one per stage, status+owner+dates)","n"),
 ("content_item ─< slot (studio+date+time+trainer; the Shooting operation)","n"),
 ("content_item ─< video_version (file link + duration + quality)","n"),
 ("video_version ─< editing_task (editor)   video_version ─< review (Shooting + Final gates)","n"),
 ("dashboard_kpis = COMPUTED from the tables — never typed by hand.","n"),
 ("","n"),
 ("CONVENTIONS / LEGEND","sub"),
 ("• Blue-filled column = primary ID (key). Grey-filled column = foreign key (points to another sheet's ID).","n"),
 ("• Status columns have dropdowns (one canonical list per process — no synonyms).","n"),
 ("• Dates as real dates (YYYY-MM-DD). Times as start/end. Studio/Trainer referenced by ID, not name.","n"),
 ("• % completion is calculated = Σ(stage.weight where status=Completed), rolled up the tree.","n"),
 ("","n"),
 ("TO POPULATE NEXT","sub"),
 ("• Add the other ~99 subjects under 'subject' (then their chapters/topics/subtopics).","n"),
 ("• Confirm stage weights incl. the new 'Shooting Review' gate (see ref_stage).","n"),
 ("• Fill real emails in 'person'.","n"),
]
ws.column_dimensions["A"].width=120
for i,(t,kind) in enumerate(lines,1):
    c=ws.cell(i,1,t)
    c.font = {"title":TITLE_FONT,"sub":SUB_FONT}.get(kind, BASE_FONT)
ws.sheet_view.showGridLines=False

# reference
write_table(wb,"ref_studio",["studio_id","studio_name","active"],
            [["STU1","Studio 1","Yes"],["STU2","Studio 2","Yes"],["STU3","Studio 3","Yes"],["STU4","Studio 4","Yes"]],
            id_cols={"studio_id"}, validations={"active":YN})
write_table(wb,"ref_stage",["stage_id","stage_name","sequence","weight","is_gate","notes"],
            [["SG1","PPT",1,0.20,"No",""],["SG2","Script",2,0.20,"No",""],
             ["SG3","Presentation",3,0.15,"No",""],["SG4","Shooting",4,0.25,"No","the recording"],
             ["SG5","Shooting Review",5,0.00,"Yes","QC gate after shooting — set a weight if desired"],
             ["SG6","Editing",6,0.10,"No",""],["SG7","Final Review",7,0.10,"Yes","final approval gate"]],
            id_cols={"stage_id"}, validations={"is_gate":YN},
            note="Weights currently sum to 1.00 (Shooting Review = gate, weight 0). Re-distribute if the QC step should carry weight.",
            widths={"notes":42,"stage_name":18})
write_table(wb,"ref_status",["process","allowed_value"],
            [["stage",v] for v in STAGE_STATUS.split(",")] +
            [["slot",v] for v in SLOT_STATUS.split(",")] +
            [["recording",v] for v in REC_STATUS.split(",")] +
            [["editing",v] for v in EDIT_STATUS.split(",")] +
            [["review",v] for v in REVIEW_STATUS.split(",")] +
            [["quality",v] for v in QUALITY.split(",")],
            widths={"allowed_value":22})

# people
write_table(wb,"person",["person_id","full_name","email","role","trainer_type","active"],
            [["P001","Madhu","madhu@example.com","trainer","Internal","Yes"],
             ["P002","Akash","akash@example.com","trainer","Internal","Yes"],
             ["P003","Leo","leo@example.com","trainer","Internal","Yes"],
             ["P004","Amy","amy@example.com","trainer","Internal","Yes"],
             ["P005","Pamy","pamy@example.com","trainer","External","Yes"],
             ["P010","Editor One","editor1@example.com","editor","","Yes"],
             ["P020","Reviewer One","review1@example.com","reviewer","","Yes"],
             ["P900","Priya (Admin)","priya@example.com","admin","","Yes"],
             ["P800","Manager","manager@example.com","manager","","Yes"]],
            id_cols={"person_id"}, validations={"role":ROLE,"trainer_type":TTYPE,"active":YN},
            widths={"email":26,"full_name":16})

# curriculum
write_table(wb,"main_subject",["main_subject_id","main_subject_name"], main_rows,
            id_cols={"main_subject_id"}, widths={"main_subject_name":24})
write_table(wb,"subject",["subject_id","main_subject_id","subject_name","sequence"], subj_rows,
            id_cols={"subject_id"}, fk_cols={"main_subject_id"},
            note="Add the other ~99 subjects here (Python, Playwright, DevOps, …), each with its own subject_id.",
            widths={"subject_name":22})
write_table(wb,"chapter",["chapter_id","subject_id","chapter_name","sequence"], chap_rows,
            id_cols={"chapter_id"}, fk_cols={"subject_id"}, widths={"chapter_name":24})
write_table(wb,"topic",["topic_id","chapter_id","topic_name","sequence"], top_rows,
            id_cols={"topic_id"}, fk_cols={"chapter_id"}, widths={"topic_name":22})
write_table(wb,"subtopic",["subtopic_id","topic_id","subtopic_name","sequence"], st_rows,
            id_cols={"subtopic_id"}, fk_cols={"topic_id"}, widths={"subtopic_name":34})

# spine
write_table(wb,"content_item",["content_item_id","subtopic_id","planned_trainer_id","target_date","notes"], ci_rows,
            id_cols={"content_item_id"}, fk_cols={"subtopic_id","planned_trainer_id"},
            note="One row per sub-topic = the unit that flows through all 7 stages.",
            widths={"notes":24})

# pipeline progress — example for first content item across 7 stages
ex_ci = ci_rows[0][0] if ci_rows else "CI00001"
istage = [[f"IS{ i:05d}", ex_ci, sg, "Not Started", "", None, None, ""]
          for i,sg in enumerate(["SG1","SG2","SG3","SG4","SG5","SG6","SG7"],1)]
istage[0][3]="Completed"; istage[0][4]="P001"
istage[3][3]="In Progress"; istage[3][4]="P001"
write_table(wb,"item_stage",
            ["item_stage_id","content_item_id","stage_id","status","owner_id","started_on","completed_on","remarks"],
            istage, id_cols={"item_stage_id"}, fk_cols={"content_item_id","stage_id","owner_id"},
            validations={"status":STAGE_STATUS},
            note="EXAMPLE rows for one content item. App auto-creates 7 rows (one per stage) per content item.",
            widths={"remarks":24})

# scheduling — from the Slots examples
write_table(wb,"slot",
            ["slot_id","date","start_time","end_time","studio_id","trainer_id","content_item_id",
             "slot_status","recording_status","reason_missed","revised_date"],
            [["SL00001","2026-06-03","10:30","12:30","STU1","P002",ex_ci,"Completed","Completed","",None],
             ["SL00002","2026-06-04","14:30","16:30","STU2","P003","CI00002","Scheduled","Not Started","",None],
             ["SL00003","2026-06-05","12:30","14:30","STU3","P004","CI00003","Missed","Not Started","Trainer on leave","2026-06-09"]],
            id_cols={"slot_id"}, fk_cols={"studio_id","trainer_id","content_item_id"},
            validations={"slot_status":SLOT_STATUS,"recording_status":REC_STATUS},
            widths={"reason_missed":18})

# outputs
write_table(wb,"video_version",
            ["video_id","content_item_id","slot_id","version_no","file_link","duration_min","quality_rating","recorded_on"],
            [["VID00001",ex_ci,"SL00001",1,"gdrive://.../coreJava_print_v1.mp4",48,"Good","2026-06-03"]],
            id_cols={"video_id"}, fk_cols={"content_item_id","slot_id"},
            validations={"quality_rating":QUALITY}, widths={"file_link":34})
write_table(wb,"editing_task",
            ["edit_id","video_id","editor_id","status","started_on","completed_on"],
            [["ED00001","VID00001","P010","In Progress","2026-06-04",None]],
            id_cols={"edit_id"}, fk_cols={"video_id","editor_id"}, validations={"status":EDIT_STATUS})
write_table(wb,"review",
            ["review_id","video_id","review_type","reviewer_id","status","decided_on","remarks"],
            [["RV00001","VID00001","Shooting Review","P020","Approved","2026-06-03","audio ok"],
             ["RV00002","VID00001","Final Review","P020","Pending","",""]],
            id_cols={"review_id"}, fk_cols={"video_id","reviewer_id"},
            validations={"review_type":REVIEW_TYPE,"status":REVIEW_STATUS}, widths={"remarks":18})

# dashboard kpis (documentation only)
write_table(wb,"dashboard_kpis",["kpi","level","how_it_is_computed"],
            [["Overall completion %","program","weighted avg of content_item stage-weights completed"],
             ["Subject completion %","subject","avg completion of its content_items"],
             ["Shooting completion %","stage","count(item_stage stage=Shooting, status=Completed)/total"],
             ["Pending recordings","ops","count(content_item with Shooting != Completed)"],
             ["Missed slots","ops","count(slot.slot_status = Missed)"],
             ["Videos pending editing","ops","count(video_version with no Completed editing_task)"],
             ["Videos pending review","ops","count(review.status = Pending)"],
             ["Room utilisation %","ops","booked slot-hours / available slot-hours per studio"],
             ["Trainer productivity %","people","completed shootings by trainer / assigned"],
             ["Quality score","quality","avg(quality_rating mapped Excellent=4..Poor=1)"]],
            widths={"how_it_is_computed":58,"kpi":24})

wb.save(OUT)
# verify
from openpyxl import load_workbook
wb2 = load_workbook(OUT)
print("SAVED:", OUT)
print("counts -> main:%d subject:%d chapter:%d topic:%d subtopic:%d content_item:%d"
      % (len(main_rows),len(subj_rows),len(chap_rows),len(top_rows),len(st_rows),len(ci_rows)))
print("sheets:", wb2.sheetnames)
