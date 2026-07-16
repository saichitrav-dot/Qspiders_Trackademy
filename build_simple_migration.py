import sys, pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding="utf-8")

SRC = r"C:\Users\praks\OneDrive\Desktop\Rec app.xlsx"
OUT = r"C:\Users\praks\Music\RecTrack-UX\RecTrack_Migration_Simple.xlsx"

STAGE = "Not Started,In Progress,Blocked,Completed"
ROLE  = "Program Head,Trainer,Editor,Reviewer,Manager"
TTYPE = "Internal,External"
STUDIO= "Studio 1,Studio 2,Studio 3,Studio 4"
SLOTST= "Scheduled,In Progress,Completed,Missed,Rescheduled"
YN    = "Yes,No"

ARIAL="Arial"
HF=PatternFill("solid",fgColor="3B3F8C"); HINT=PatternFill("solid",fgColor="Eef0FF")
KEYINFO=PatternFill("solid",fgColor="E6F6EC")
HFONT=Font(name=ARIAL,bold=True,color="FFFFFF",size=10); BASE=Font(name=ARIAL,size=10)
TITLE=Font(name=ARIAL,bold=True,size=15,color="1F2340"); SUB=Font(name=ARIAL,bold=True,size=11,color="3B3F8C")
thin=Side(style="thin",color="DADDE3"); BD=Border(left=thin,right=thin,top=thin,bottom=thin)

def sheet(wb,name,headers,rows,widths=None,dvs=None,hint=None):
    ws=wb.create_sheet(name); r0=1
    if hint:
        c=ws.cell(1,1,hint); c.font=Font(name=ARIAL,italic=True,size=9,color="3B3F8C"); c.fill=HINT; r0=2
    for j,h in enumerate(headers,1):
        c=ws.cell(r0,j,h); c.font=HFONT; c.fill=HF; c.alignment=Alignment("left","center"); c.border=BD
    for i,row in enumerate(rows,r0+1):
        for j,val in enumerate(row,1):
            c=ws.cell(i,j,val); c.font=BASE; c.border=BD; c.alignment=Alignment(vertical="center")
    for j,h in enumerate(headers,1):
        ws.column_dimensions[get_column_letter(j)].width=(widths or {}).get(h,max(11,min(34,len(str(h))+3)))
    ws.freeze_panes=ws.cell(r0+1,1); ws.row_dimensions[r0].height=20
    if dvs:
        for h,lst in dvs.items():
            if h in headers:
                col=get_column_letter(headers.index(h)+1)
                dv=DataValidation(type="list",formula1=f'"{lst}"',allow_blank=True)
                ws.add_data_validation(dv); dv.add(f"{col}{r0+1}:{col}{r0+3000}")
    return ws

# ---- read your Excel, clean hierarchy, one row per sub-topic ----
df=pd.read_excel(SRC,sheet_name="Subjects and Status")
df.columns=[str(c).strip() for c in df.columns]
for c in ["Main Subject","Subject","ChapterName","TopicName"]: df[c]=df[c].ffill()
df["TopicName"]=df["TopicName"].fillna("General")
df=df[df["Sub-topics"].notna()].copy().reset_index(drop=True)

main_rows=[]
for idx,r in df.iterrows():
    first = idx==0
    main_rows.append([
        r["Main Subject"], r["Subject"], r["ChapterName"], r["TopicName"], str(r["Sub-topics"]).strip(),
        "Madhu" if first else "",            # planned trainer (example on row 1)
        "Internal",                           # trainer type (from your file)
        "Completed" if first else "",         # PPT
        "Not Started" if first else "",       # Script
        "In Progress" if first else "",       # Presentation
        "" , "", "", "",                      # Shooting, Shoot Review, Editing, Final Review (blank to fill)
        "", ""                                # target date, notes
    ])

wb=Workbook(); wb.remove(wb.active)

# Read me
ws=wb.create_sheet("Read me"); ws.column_dimensions["A"].width=118; ws.sheet_view.showGridLines=False
for i,(t,k) in enumerate([
  ("RecTrack — Simple Migration File","title"),
  ("Just 3 working tabs. Fill these in plain language — no ID numbers, ever.","sub"),
  ("",""),
  ("• Tab 'Curriculum + Status' = the main one. One row per sub-topic. Pick a status from the dropdown for each stage.","n"),
  ("• Tab 'People' = your team and their role.   • Tab 'Schedule' = recording slots (or manage these in the app later).","n"),
  ("",""),
  ("WHY NO ID COLUMNS?  The system creates the keys (primary/foreign) automatically when this is imported —","keyinfo"),
  ("you never type one. That is exactly why the 'duplicate IDs from copies' problem can't happen again.","keyinfo"),
  ("",""),
  ("Core Java is already filled in from your file as the example. Add the other subjects the same way.","n"),
], 1):
    c=ws.cell(i,1,t); c.font={"title":TITLE,"sub":SUB,"keyinfo":Font(name=ARIAL,size=10,color="0F7A37",bold=True)}.get(k,BASE)
    if k=="keyinfo": c.fill=KEYINFO

# 1. Curriculum + Status  (wide & familiar, like your original)
sheet(wb,"Curriculum + Status",
   ["Main Subject","Subject","Chapter","Topic","Sub-topic","Planned Trainer","Trainer Type",
    "PPT","Script","Presentation","Shooting","Shooting Review","Editing","Final Review","Target Date","Notes"],
   main_rows,
   widths={"Main Subject":15,"Subject":13,"Chapter":20,"Topic":16,"Sub-topic":30,"Planned Trainer":14,
           "Trainer Type":12,"Target Date":12,"Notes":20},
   dvs={"Trainer Type":TTYPE,"PPT":STAGE,"Script":STAGE,"Presentation":STAGE,"Shooting":STAGE,
        "Shooting Review":STAGE,"Editing":STAGE,"Final Review":STAGE},
   hint="One row per sub-topic. Pick each stage's status from the dropdown. Hierarchy is filled down — no blank 'same as above' cells.")

# 2. People
sheet(wb,"People",
   ["Name","Email","Role","Internal/External","Active"],
   [["Madhu","madhu@example.com","Trainer","Internal","Yes"],
    ["Akash","akash@example.com","Trainer","Internal","Yes"],
    ["Leo","leo@example.com","Trainer","Internal","Yes"],
    ["Amy","amy@example.com","Trainer","Internal","Yes"],
    ["Pamy","pamy@example.com","Trainer","External","Yes"],
    ["Priya","priya@example.com","Program Head","","Yes"],
    ["(add editor later)","","Editor","","Yes"],
    ["(add reviewer later)","","Reviewer","","Yes"]],
   widths={"Email":26,"Name":18,"Internal/External":15},
   dvs={"Role":ROLE,"Internal/External":TTYPE,"Active":YN},
   hint="Phase 1 = Program Head + Trainers. Add Editor / Reviewer / Manager when you switch them on.")

# 3. Schedule (Slots)
sheet(wb,"Schedule",
   ["Date","Start","End","Studio","Trainer","Sub-topic","Status","Notes"],
   [["2026-06-03","10:30","12:30","Studio 1","Akash","Print","Completed",""],
    ["2026-06-03","12:30","14:30","Studio 2","Leo","Operators","In Progress",""],
    ["2026-06-04","08:30","10:30","Studio 3","Amy","Inheritance","Scheduled",""],
    ["2026-06-05","12:30","14:30","Studio 4","Pamy","Arrays","Missed","trainer on leave"]],
   widths={"Sub-topic":22,"Notes":22},
   dvs={"Studio":STUDIO,"Status":SLOTST},
   hint="Example slots. Going forward these are created in the app — this tab is just for the initial load.")

wb.save(OUT)
print("SAVED:",OUT)
print("sub-topic rows migrated:",len(main_rows))
print("tabs:",wb.sheetnames)
