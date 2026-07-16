import sys, pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding="utf-8")

SRC = r"C:\Users\praks\OneDrive\Desktop\Rec app.xlsx"
OUT = r"C:\Users\praks\Music\RecTrack-UX\RecTrack_DataModel.xlsx"

STAGE_STATUS = "Not Started,In Progress,Blocked,Completed"
SLOT_STATUS  = "Scheduled,In Progress,Completed,Missed,Rescheduled,Swapped"
REC_STATUS   = "Not Started,Recording,Completed"
EDIT_STATUS  = "Not Started,In Progress,Rework,Completed"
REVIEW_STATUS= "Pending,Approved,Rejected,Rework"
QUALITY      = "Excellent,Good,Average,Poor"
ROLE         = "admin,trainer,editor,reviewer,manager"
TTYPE        = "Internal,External"
YN           = "Yes,No"
REVIEW_TYPE  = "Shooting Review,Final Review"

ARIAL="Arial"
HDR_FILL=PatternFill("solid",fgColor="3B3F8C"); ID_FILL=PatternFill("solid",fgColor="E8EAF6")
FK_FILL=PatternFill("solid",fgColor="F1F3F5"); NEW_FILL=PatternFill("solid",fgColor="E6F6EC")
NOTE_FILL=PatternFill("solid",fgColor="FFF7E0")
HDR_FONT=Font(name=ARIAL,bold=True,color="FFFFFF",size=10); BASE=Font(name=ARIAL,size=10)
TITLE=Font(name=ARIAL,bold=True,size=14,color="1F2340"); SUB=Font(name=ARIAL,bold=True,size=11,color="3B3F8C")
thin=Side(style="thin",color="D9DCE3"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

def wt(wb,name,headers,rows,id_cols=(),fk_cols=(),new_cols=(),widths=None,validations=None,note=None):
    ws=wb.create_sheet(name); r0=1
    if note:
        c=ws.cell(1,1,note); c.font=Font(name=ARIAL,italic=True,size=9,color="8A6D00"); c.fill=NOTE_FILL; r0=2
    for j,h in enumerate(headers,1):
        c=ws.cell(r0,j,h); c.font=HDR_FONT; c.fill=HDR_FILL; c.alignment=Alignment("left","center"); c.border=BORDER
    for i,row in enumerate(rows,r0+1):
        for j,val in enumerate(row,1):
            c=ws.cell(i,j,val); c.font=BASE; c.border=BORDER; c.alignment=Alignment(vertical="center")
            h=headers[j-1]
            if h in id_cols: c.fill=ID_FILL
            elif h in fk_cols: c.fill=FK_FILL
            elif h in new_cols: c.fill=NEW_FILL
    for j,h in enumerate(headers,1):
        ws.column_dimensions[get_column_letter(j)].width=(widths or {}).get(h,max(12,min(42,len(str(h))+4)))
    ws.freeze_panes=ws.cell(r0+1,1); ws.row_dimensions[r0].height=20
    if validations:
        for h,lst in validations.items():
            if h in headers:
                col=get_column_letter(headers.index(h)+1)
                dv=DataValidation(type="list",formula1=f'"{lst}"',allow_blank=True)
                ws.add_data_validation(dv); dv.add(f"{col}{r0+1}:{col}{r0+2000}")
    return ws

# ---- migrate Core Java ----
df=pd.read_excel(SRC,sheet_name="Subjects and Status")
df.columns=[str(c).strip() for c in df.columns]
for c in ["Main Subject","Subject","ChapterName","TopicName"]: df[c]=df[c].ffill()
df["TopicName"]=df["TopicName"].fillna("General")
df=df[df["Sub-topics"].notna()].copy().reset_index(drop=True)
main_rows,subj_rows,chap_rows,top_rows,st_rows,ci_rows=[],[],[],[],[],[]
mmap,smap,cmap,tmap={},{},{},{}; mc=sc=cc=tc=stc=0
for _,r in df.iterrows():
    ms,su,ch,tp,sb=r["Main Subject"],r["Subject"],r["ChapterName"],r["TopicName"],str(r["Sub-topics"]).strip()
    if ms not in mmap: mc+=1; mmap[ms]=f"MS{mc:02d}"; main_rows.append([mmap[ms],ms])
    mid=mmap[ms]
    if (mid,su) not in smap: sc+=1; smap[(mid,su)]=f"SUB{sc:03d}"; subj_rows.append([smap[(mid,su)],mid,su,sc])
    sid=smap[(mid,su)]
    if (sid,ch) not in cmap: cc+=1; cmap[(sid,ch)]=f"CH{cc:04d}"; chap_rows.append([cmap[(sid,ch)],sid,ch,len(chap_rows)+1])
    cid=cmap[(sid,ch)]
    if (cid,tp) not in tmap: tc+=1; tmap[(cid,tp)]=f"TP{tc:04d}"; top_rows.append([tmap[(cid,tp)],cid,tp,len(top_rows)+1])
    tid=tmap[(cid,tp)]
    stc+=1; stid=f"ST{stc:05d}"; seq=sum(1 for x in st_rows if x[1]==tid)+1
    st_rows.append([stid,tid,sb,seq]); ci_rows.append([f"CI{stc:05d}",stid,"P001",None,""])

wb=Workbook(); wb.remove(wb.active)

# README
ws=wb.create_sheet("_README"); ws.column_dimensions["A"].width=122; ws.sheet_view.showGridLines=False
lines=[
 ("RecTrack — Data Model & Migration Template (v2)","title"),
 ("Target build: managed cloud (Supabase Postgres) + web app. This workbook is the ONE-TIME migration source.","sub"),
 ("","n"),
 ("LOCKED DESIGN","sub"),
 ("• Deliverable = one content_item per Sub-topic.  • Pipeline = 7 stages with 2 gates (Shooting Review + Final Review).","n"),
 ("• 2-day turnaround WINDOW per stage — see ref_config.sla_window_days and the due_on / breached / delay_days fields.","n"),
 ("• Roles: launch Phase 1 = Program Head + Trainer; Phase 2 = Editing team + Reviewer + Management (person.launch_phase).","n"),
 ("• Keys here (MS01, ST00001…) are READABLE for migration only. On import the system assigns permanent UUIDs — so the","n"),
 ("  'duplicate IDs from copies' problem cannot happen (one cloud DB, auto-generated keys, no copies).","n"),
 ("","n"),
 ("ENTITY MAP","sub"),
 ("main_subject → subject → chapter → topic → subtopic → content_item (the spine)","n"),
 ("content_item ─< item_stage (7 rows: status, owner, due_on, breached, delay_days)","n"),
 ("content_item ─< slot (studio+date+time+trainer = the Shooting op)   content_item ─< video_version (link+quality)","n"),
 ("video_version ─< editing_task (editor)    video_version ─< review (Shooting + Final gates)","n"),
 ("dashboard_kpis = COMPUTED from the tables — never typed.","n"),
 ("","n"),
 ("HOW TO FILL (migration order)","sub"),
 ("1) ref_* sheets are pre-filled (studios, stages, statuses, config) — adjust if needed.","n"),
 ("2) person — add your people; set role + launch_phase + real emails.","n"),
 ("3) Curriculum — Core Java is migrated as the example. Add the other ~99 subjects (subject → chapter → topic → subtopic).","n"),
 ("4) content_item auto = one per subtopic. slot / item_stage / video_version / editing_task / review fill as work happens.","n"),
 ("","n"),
 ("LEGEND","sub"),
 ("Blue = primary ID (key).  Grey = foreign key.  Green = NEW v2 field (SLA window / delay tracking).  Yellow = note.","n"),
]
for i,(t,k) in enumerate(lines,1):
    c=ws.cell(i,1,t); c.font={"title":TITLE,"sub":SUB}.get(k,BASE)

# ref_config (NEW)
wt(wb,"ref_config",["setting","value","notes"],
   [["sla_window_days",2,"turnaround window allowed per stage (the 2-day rule)"],
    ["workday_start","08:30","first slot start"],["workday_end","20:30","last slot end"],
    ["slot_length_hours",2,"length of one recording slot"],["studios_count",4,"Studio 1-4"],
    ["on_window_target",0.85,"min % of items each person finishes inside window"],
    ["rework_target",0.05,"max acceptable rework rate"],["reminder_lead_min",15,"pre-slot reminder lead time"],
    ["notify_channel","email","WhatsApp deferred to v3"]],
   id_cols={"setting"}, widths={"notes":52,"setting":20}, note="Global rules as DATA (config, not code) — change here, no developer needed.")

wt(wb,"ref_studio",["studio_id","studio_name","active"],
   [["STU1","Studio 1","Yes"],["STU2","Studio 2","Yes"],["STU3","Studio 3","Yes"],["STU4","Studio 4","Yes"]],
   id_cols={"studio_id"}, validations={"active":YN})

wt(wb,"ref_stage",["stage_id","stage_name","sequence","weight","is_gate","sla_window_days","notes"],
   [["SG1","PPT",1,0.20,"No",2,""],["SG2","Script",2,0.20,"No",2,""],["SG3","Presentation",3,0.15,"No",2,""],
    ["SG4","Shooting",4,0.25,"No",2,"the recording"],
    ["SG5","Shooting Review",5,0.00,"Yes",1,"GATE 1 — QC right after shooting"],
    ["SG6","Editing",6,0.10,"No",2,""],
    ["SG7","Final Review",7,0.10,"Yes",1,"GATE 2 — final approval"]],
   id_cols={"stage_id"}, new_cols={"sla_window_days"}, validations={"is_gate":YN},
   note="7 stages, 2 gates. Weights sum to 1.00. sla_window_days = the 2-day window per stage (gates default 1).",
   widths={"notes":34,"stage_name":18})

wt(wb,"ref_status",["process","allowed_value"],
   [["stage",v] for v in STAGE_STATUS.split(",")]+[["slot",v] for v in SLOT_STATUS.split(",")]+
   [["recording",v] for v in REC_STATUS.split(",")]+[["editing",v] for v in EDIT_STATUS.split(",")]+
   [["review",v] for v in REVIEW_STATUS.split(",")]+[["quality",v] for v in QUALITY.split(",")],
   widths={"allowed_value":22})

wt(wb,"person",["person_id","full_name","email","role","trainer_type","launch_phase","active"],
   [["P001","Madhu","madhu@example.com","trainer","Internal",1,"Yes"],
    ["P002","Akash","akash@example.com","trainer","Internal",1,"Yes"],
    ["P003","Leo","leo@example.com","trainer","Internal",1,"Yes"],
    ["P004","Amy","amy@example.com","trainer","Internal",1,"Yes"],
    ["P005","Pamy","pamy@example.com","trainer","External",1,"Yes"],
    ["P900","Priya (Program Head)","priya@example.com","admin","",1,"Yes"],
    ["P010","Editor One","editor1@example.com","editor","",2,"Yes"],
    ["P020","Reviewer One","review1@example.com","reviewer","",2,"Yes"],
    ["P800","Manager","manager@example.com","manager","",2,"Yes"]],
   id_cols={"person_id"}, new_cols={"launch_phase"}, validations={"role":ROLE,"trainer_type":TTYPE,"active":YN},
   note="launch_phase 1 = live first (Program Head + Trainers). launch_phase 2 = switched on later.",
   widths={"email":26,"full_name":20})

wt(wb,"main_subject",["main_subject_id","main_subject_name"],main_rows,id_cols={"main_subject_id"},widths={"main_subject_name":24})
wt(wb,"subject",["subject_id","main_subject_id","subject_name","sequence"],subj_rows,
   id_cols={"subject_id"},fk_cols={"main_subject_id"},widths={"subject_name":22},
   note="Add the other ~99 subjects here (Advanced Java, Spring Boot, Selenium…), each its own subject_id.")
wt(wb,"chapter",["chapter_id","subject_id","chapter_name","sequence"],chap_rows,id_cols={"chapter_id"},fk_cols={"subject_id"},widths={"chapter_name":24})
wt(wb,"topic",["topic_id","chapter_id","topic_name","sequence"],top_rows,id_cols={"topic_id"},fk_cols={"chapter_id"},widths={"topic_name":22})
wt(wb,"subtopic",["subtopic_id","topic_id","subtopic_name","sequence"],st_rows,id_cols={"subtopic_id"},fk_cols={"topic_id"},widths={"subtopic_name":34})

wt(wb,"content_item",["content_item_id","subtopic_id","planned_trainer_id","target_date","notes"],ci_rows,
   id_cols={"content_item_id"},fk_cols={"subtopic_id","planned_trainer_id"},widths={"notes":22},
   note="One row per sub-topic = the unit that flows through all 7 stages.")

ex=ci_rows[0][0] if ci_rows else "CI00001"
istage=[[f"IS{i:05d}",ex,sg,"Not Started","",None,None,None,"No",0,""] for i,sg in enumerate(["SG1","SG2","SG3","SG4","SG5","SG6","SG7"],1)]
istage[0][3]="Completed"; istage[0][4]="P001"
istage[3][3]="In Progress"; istage[3][4]="P001"
istage[5][3]="Blocked"; istage[5][4]="P010"; istage[5][8]="Yes"; istage[5][9]=1  # editing breached, 1 delay day
wt(wb,"item_stage",
   ["item_stage_id","content_item_id","stage_id","status","owner_id","started_on","due_on","completed_on","breached","delay_days","remarks"],
   istage, id_cols={"item_stage_id"}, fk_cols={"content_item_id","stage_id","owner_id"},
   new_cols={"due_on","breached","delay_days"}, validations={"status":STAGE_STATUS,"breached":YN},
   note="EXAMPLE rows for one item. App auto-creates 7 rows/item. due_on = started_on + stage SLA window; breached + delay_days track the 2-day rule.",
   widths={"remarks":20})

wt(wb,"slot",
   ["slot_id","date","start_time","end_time","studio_id","trainer_id","content_item_id","slot_status","recording_status","reason_missed","revised_date","delay_days"],
   [["SL00001","2026-06-03","10:30","12:30","STU1","P002",ex,"Completed","Completed","",None,0],
    ["SL00002","2026-06-04","14:30","16:30","STU2","P003","CI00002","Scheduled","Not Started","",None,0],
    ["SL00003","2026-06-05","12:30","14:30","STU3","P004","CI00003","Missed","Not Started","Trainer on leave","2026-06-06",1]],
   id_cols={"slot_id"}, fk_cols={"studio_id","trainer_id","content_item_id"}, new_cols={"delay_days"},
   validations={"slot_status":SLOT_STATUS,"recording_status":REC_STATUS}, widths={"reason_missed":18})

wt(wb,"video_version",
   ["video_id","content_item_id","slot_id","version_no","file_link","duration_min","quality_rating","recorded_on"],
   [["VID00001",ex,"SL00001",1,"gdrive://core-java/operators_v1.mp4",48,"Good","2026-06-03"]],
   id_cols={"video_id"}, fk_cols={"content_item_id","slot_id"}, validations={"quality_rating":QUALITY}, widths={"file_link":34})

wt(wb,"editing_task",["edit_id","video_id","editor_id","status","started_on","due_on","completed_on"],
   [["ED00001","VID00001","P010","In Progress","2026-06-04","2026-06-06",None]],
   id_cols={"edit_id"}, fk_cols={"video_id","editor_id"}, new_cols={"due_on"}, validations={"status":EDIT_STATUS})

wt(wb,"review",["review_id","video_id","review_type","reviewer_id","status","decided_on","remarks"],
   [["RV00001","VID00001","Shooting Review","P020","Approved","2026-06-03","audio ok"],
    ["RV00002","VID00001","Final Review","P020","Pending","",""]],
   id_cols={"review_id"}, fk_cols={"video_id","reviewer_id"}, validations={"review_type":REVIEW_TYPE,"status":REVIEW_STATUS}, widths={"remarks":18})

wt(wb,"dashboard_kpis",["kpi","level","how_it_is_computed"],
   [["Program completion %","program","weighted avg of content_item stage-weights completed"],
    ["Subject completion %","subject","avg completion of its content_items"],
    ["Forecast finish date","program","remaining stage-points ÷ throughput (per week)"],
    ["Throughput / week","program","content_items advancing per week"],
    ["On-window SLA %","ops","item_stages completed with breached=No ÷ total"],
    ["Breached / due-today","ops","count item_stage where due_on ≤ today and status ≠ Completed"],
    ["Delay days (total)","ops","sum(item_stage.delay_days)+sum(slot.delay_days)"],
    ["Missed slots","ops","count(slot.slot_status = Missed)"],
    ["Videos pending editing","ops","video_version with no Completed editing_task"],
    ["Videos pending review","ops","count(review.status = Pending)"],
    ["Studio utilisation %","ops","booked slot-hours ÷ available slot-hours"],
    ["Trainer/editor productivity","people","completed items by person ÷ assigned"],
    ["Quality score","quality","avg(quality_rating Excellent=4..Poor=1)"]],
   widths={"how_it_is_computed":56,"kpi":24})

wb.save(OUT)
from openpyxl import load_workbook
wb2=load_workbook(OUT)
print("SAVED:",OUT)
print("curriculum -> main %d / subject %d / chapter %d / topic %d / subtopic %d / content_item %d"
      %(len(main_rows),len(subj_rows),len(chap_rows),len(top_rows),len(st_rows),len(ci_rows)))
print("sheets:",wb2.sheetnames)
